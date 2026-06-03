import streamlit as st
from openpyxl import Workbook
from io import BytesIO

# ----------------------------------------
# TAX CALCULATION
# ----------------------------------------
def calculate_proprietorship_tax(profit):
    tax = 0

    if profit <= 500000:
        tax = 0

    elif profit <= 700000:
        tax = (profit - 500000) * 0.10

    elif profit <= 1000000:
        tax = (200000 * 0.10) + (profit - 700000) * 0.20

    elif profit <= 2000000:
        tax = (200000 * 0.10) + (300000 * 0.20) + (profit - 1000000) * 0.30

    elif profit <= 5000000:
        tax = (
            (200000 * 0.10)
            + (300000 * 0.20)
            + (1000000 * 0.30)
            + (profit - 2000000) * 0.36
        )

    else:
        tax = (
            (200000 * 0.10)
            + (300000 * 0.20)
            + (1000000 * 0.30)
            + (3000000 * 0.36)
            + (profit - 5000000) * 0.39
        )

    return tax


st.set_page_config(page_title="Advance Tax Calculator")

st.title("Firm Tax Installment Calculator")

firm_name = st.text_input("Firm Name")

sales = st.number_input(
    "Net Sales Amount",
    min_value=0.0,
    value=0.0
)

np_percent = st.number_input(
    "Net Profit %",
    min_value=0.0,
    value=0.0
)

firm_type = st.selectbox(
    "Firm Type",
    ["Proprietorship", "Others"]
)

installment = st.selectbox(
    "Installment",
    ["Up to Poush", "Up to Chaitra", "Up to Ashad"]
)

if st.button("Calculate Tax"):

    profit = sales * np_percent / 100

    if firm_type == "Proprietorship":
        tax = calculate_proprietorship_tax(profit)
    else:
        tax = profit * 0.25

    if installment == "Up to Poush":
        liability = tax * 0.40
    elif installment == "Up to Chaitra":
        liability = tax * 0.70
    else:
        liability = tax

    st.success("Calculation Completed")

    st.write("### Result")

    st.write(f"Firm Name: {firm_name}")
    st.write(f"Net Sales: {sales:,.2f}")
    st.write(f"Net Profit %: {np_percent:.2f}%")
    st.write(f"Net Profit: {profit:,.2f}")
    st.write(f"Firm Type: {firm_type}")
    st.write(f"Total Tax: {tax:,.2f}")
    st.write(f"Installment: {installment}")
    st.write(f"Tax Liability: {liability:,.2f}")

    # Excel Export

    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Firm Tax Installment Report"

    data = {
        "Firm Name": firm_name,
        "Net Sales": sales,
        "Net Profit %": np_percent,
        "Net Profit": profit,
        "Firm Type": firm_type,
        "Total Tax": tax,
        "Installment": installment,
        "Tax Liability": liability
    }

    row = 3

    for key, value in data.items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1

    excel_file = BytesIO()
    wb.save(excel_file)

    st.download_button(
        label="Download Excel Report",
        data=excel_file.getvalue(),
        file_name="Tax_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )